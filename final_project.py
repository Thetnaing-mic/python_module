import os
import pyodbc
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook , load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

load_dotenv()

SERVER = os.getenv('SERVER')
DATABASE = os.getenv('DATABASE')
USER = os.getenv('DB_USER')
PASSWORD = os.getenv('DB_PASSWORD')

connection_str = f'DRIVER={{SQL Server}}; SERVER={SERVER}; DATABASE={DATABASE}; UID={USER}; PWD={PASSWORD};'
print(connection_str)

try:
    conn = pyodbc.connect(connection_str)
    cursor = conn.cursor()
    print("Connection established successfully.")
except pyodbc.Error as ex:
    sqlstate = ex.args[1]
    print(f"An error occurred: {sqlstate}")

# 2
# cursor.execute('''CREATE TABLE my_newmovie (
#     ID INT PRIMARY KEY IDENTITY(1, 1),
#     Release_Date DATE,
#     Title VARCHAR(255),
#     Overview VARCHAR(2000),
#     Popularity FLOAT,
#     Vote_Count INT,
#     Vote_Average FLOAT,
#     Original_Language VARCHAR(2),
#     Genre VARCHAR(255),
#     Poster_Url VARCHAR(255)
# )''')


# # 3
# df = pd.read_csv("./mymoviedb.csv", lineterminator='\n')
# df.dropna(how = "all")
# print(df)

# # 4
# insert_query = '''
# INSERT INTO my_movie (Release_Date, Title, Overview, Popularity, Vote_Count, Vote_Average, Original_Language, Genre, Poster_Url)
# VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
# '''
# for index, row in df.iterrows():
#     cursor.execute(insert_query, 
#                     row['Release_Date'], 
#                     row['Title'], 
#                     row['Overview'], 
#                     row['Popularity'],
#                     row['Vote_Count'],
#                     row['Vote_Average'],
#                     row['Original_Language'], 
#                     row['Genre'], 
#                     row['Poster_Url'])
# conn.commit()

# # 5
# df = pd.read_sql("SELECT * FROM my_movie", conn)
# print(df)

# # 6
# language_count_query = "SELECT Original_Language, COUNT(*) AS Movie FROM my_movie GROUP BY Original_Language"

# # 7
# genre_rating_query = "SELECT Genre, AVG(Vote_Average) AS Average_Rating FROM my_movie GROUP BY Genre"

# # 8

# language_count = pd.read_sql(language_count_query, conn)
# # print(language_count)
# genre_rating = pd.read_sql(genre_rating_query, conn)
# # print(genre_rating)
# wb = Workbook()

# ws1 = wb.create_sheet(title = "Language Counts")
# for r in dataframe_to_rows(language_count, index=False, header=True):
#     ws1.append(r)


# ws2 = wb.create_sheet(title="Genre Ratings")
# for r in dataframe_to_rows(genre_rating, index=False, header=True):
#     ws2.append(r)

# del wb['Sheet']

# wb.save('movies_analysis.xlsx')

# print("Analysis results saved to 'movies_analysis.xlsx'")

# cursor.close()
# conn.close()