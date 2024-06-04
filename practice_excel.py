from openpyxl import Workbook, load_workbook
# wb = Workbook()
# wb = load_workbook

wb = load_workbook("excel.xlsx")
ws = wb.create_sheet("Testsheet")

# １
count = 1
for i in range(2, 6):
    for j in range(2, 6):
        ws.cell(i, j, count)




# ws = wb.create_sheet("DataInput")

# for i in range(1, 11):
#     ws.cell(row=i, column=1, value=i)


# test_sheet = wb.create_sheet("TestSheet")


# for i in range(1, 11):
#     original_value = ws.cell(row=i, column=1).value
#     modified_value = original_value + 10
#     test_sheet.cell(row=i + 1, column=6, value=modified_value)  

# start_row, end_row = 2, 5
# start_col, end_col = 2, 5

# for i in range(start_row, end_row + 1):
#     # Loop through columns
#     for j in range(start_col, end_col + 1):
#         # Get the cell value
#         cell = ws.cell(row=i, column=j)
#         # Print the cell value
#         print(cell.value, end="\t")  
#     print()  









# print(ws["A1"].value)
# print(ws["B2"].value)
# ws = wb.create_sheet(title = "newSheet")
# print(ws.cell(2,1).value)

# ws["A1"] = 1
# ws.cell(1, 1, 1)

# # ws["B1"] = 2
# ws.cell(1, 2, 2)

# # ws["A2"] = 3
# ws.cell(2, 1, 3)

# # ws["B2"] = 4
# ws.cell(2, 2, 4)


# ws.append([1, 2, 3, 4])
# ws.append([1])
# ws.append([1, 2, 3])

# wb.save("ExcelResult_テッ.xlsx")